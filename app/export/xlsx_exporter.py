"""
XLSX exporter module - exports results as Excel file.
"""

import os
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database.repositories.link_repo import LinkRepository
from app.database.repositories.result_repo import ResultRepository
from app.database.repositories.scan_repo import ScanRepository
from app.config import config
from app.utils.logger import get_logger

logger = get_logger(__name__)


class XLSXExporter:
    """
    Exports scan results as an Excel file (XLSX).
    """
    
    @classmethod
    async def export(cls, job_id: int) -> str:
        """
        Export scan results to an XLSX file.
        
        Args:
            job_id: Scan job ID
            
        Returns:
            Path to the exported file
        """
        logger.info(f"Exporting job {job_id} as XLSX")
        
        # Get job info and statistics
        scan_repo = ScanRepository()
        result_repo = ResultRepository()
        link_repo = LinkRepository()
        
        job = await scan_repo.get_by_id(job_id)
        stats = await result_repo.get_statistics(job_id)
        
        if not job:
            raise ValueError(f"Job {job_id} not found")
        
        # Get all links
        all_links = await link_repo.get_by_job(
            scan_job_id=job_id,
            limit=10000,
            offset=0
        )
        
        # Create workbook
        wb = Workbook()
        
        # ============================================================
        # Sheet 1: Results
        # ============================================================
        ws_results = wb.active
        ws_results.title = "Results"
        
        # Set headers
        headers = [
            'URL', 'Platform', 'Link Type', 'Status',
            'Entity Title', 'Entity Username', 'Source ID', 'Message ID',
            'Created At'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, link in enumerate(all_links, 2):
            ws_results.cell(row=row, column=1, value=link.get('original_url', ''))
            ws_results.cell(row=row, column=2, value=link.get('platform', ''))
            ws_results.cell(row=row, column=3, value=link.get('link_type', ''))
            ws_results.cell(row=row, column=4, value=link.get('status', ''))
            ws_results.cell(row=row, column=5, value=link.get('entity_title', ''))
            ws_results.cell(row=row, column=6, value=link.get('entity_username', ''))
            ws_results.cell(row=row, column=7, value=link.get('source_id', ''))
            ws_results.cell(row=row, column=8, value=link.get('message_id', ''))
            ws_results.cell(row=row, column=9, value=link.get('created_at', ''))
        
        # Auto-fit columns
        for column in ws_results.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_results.column_dimensions[column_letter].width = adjusted_width
        
        # ============================================================
        # Sheet 2: Statistics
        # ============================================================
        ws_stats = wb.create_sheet("Statistics")
        
        # Style for stats
        stats_title_font = Font(bold=True, size=12)
        stats_value_font = Font(size=11)
        stats_alignment = Alignment(horizontal="left", vertical="center")
        
        # Job information
        ws_stats.cell(row=1, column=1, value="JOB INFORMATION")
        ws_stats.cell(row=1, column=1).font = stats_title_font
        
        job_data = [
            ("Job ID", job_id),
            ("Status", job.status.upper()),
            ("Scan Mode", job.scan_mode.upper()),
            ("Created At", job.created_at),
            ("Completed At", job.completed_at or "N/A"),
        ]
        
        for i, (label, value) in enumerate(job_data, 2):
            ws_stats.cell(row=i, column=1, value=f"{label}:")
            ws_stats.cell(row=i, column=2, value=value)
            ws_stats.cell(row=i, column=1).font = stats_value_font
            ws_stats.cell(row=i, column=2).font = stats_value_font
        
        # Statistics
        row_offset = len(job_data) + 3
        ws_stats.cell(row=row_offset, column=1, value="SCAN STATISTICS")
        ws_stats.cell(row=row_offset, column=1).font = stats_title_font
        
        if stats:
            stats_data = [
                ("Messages Scanned", job.messages_scanned),
                ("URLs Found", job.urls_found),
                ("Unique URLs", job.urls_unique),
                ("", ""),
                ("Telegram Links", stats.get('telegram_count', 0)),
                ("WhatsApp Links", stats.get('whatsapp_count', 0)),
                ("", ""),
                ("Groups", stats.get('group_count', 0)),
                ("Channels", stats.get('channel_count', 0)),
                ("Invites", stats.get('invite_count', 0)),
                ("", ""),
                ("Excluded Personal", stats.get('personal_count', 0)),
                ("Excluded Bots", stats.get('bot_count', 0)),
                ("Duplicates", stats.get('duplicate_count', 0)),
                ("Other Links", stats.get('other_count', 0)),
            ]
            
            for i, (label, value) in enumerate(stats_data, row_offset + 1):
                ws_stats.cell(row=i, column=1, value=f"{label}:")
                ws_stats.cell(row=i, column=2, value=value)
                ws_stats.cell(row=i, column=1).font = stats_value_font
                ws_stats.cell(row=i, column=2).font = stats_value_font
        
        # Auto-fit columns for stats sheet
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 20
        
        # Add borders to statistics
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws_stats.iter_rows(min_row=1, max_row=row_offset + len(stats_data) + 1):
            for cell in row:
                cell.border = thin_border
        
        # ============================================================
        # Save file
        # ============================================================
        filename = f"results_job_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = config.EXPORT_DIR / filename
        
        wb.save(file_path)
        logger.info(f"XLSX export completed: {file_path}")
        
        return str(file_path)