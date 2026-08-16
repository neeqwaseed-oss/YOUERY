"""
Integration tests for export functionality.
"""

import pytest
import os
import json
import csv
from pathlib import Path

from app.export.txt_exporter import TXTExporter
from app.export.csv_exporter import CSVExporter
from app.export.json_exporter import JSONExporter
from app.export.xlsx_exporter import XLSXExporter
from app.database.repositories.scan_repo import ScanRepository
from app.database.repositories.link_repo import LinkRepository
from app.database.repositories.result_repo import ResultRepository
from app.database.repositories.source_repo import SourceRepository
from app.database.repositories.user_repo import UserRepository
from app.config import config


@pytest.mark.asyncio
class TestExport:
    """Integration tests for export functionality."""

    async def setup_test_data(self):
        """Setup test data for export tests."""
        user_repo = UserRepository()
        source_repo = SourceRepository()
        scan_repo = ScanRepository()
        link_repo = LinkRepository()
        result_repo = ResultRepository()
        
        # Create user
        await user_repo.create_or_update(
            telegram_user_id=123456,
            username="testuser",
            first_name="Test User"
        )
        
        # Create source
        source = await source_repo.create(
            telegram_user_id=123456,
            telegram_chat_id=789012,
            title="Test Source",
            source_type="channel"
        )
        
        # Create scan job
        job = await scan_repo.create(
            telegram_user_id=123456,
            source_id=source.id,
            scan_mode="recent",
            limit=100
        )
        
        # Create links
        test_links = [
            {
                'scan_job_id': job.id,
                'source_id': source.id,
                'message_id': 1,
                'original_url': 'https://t.me/testgroup1',
                'normalized_url': 'https://t.me/testgroup1',
                'platform': 'telegram',
                'link_type': 'public_group',
                'status': 'valid',
                'entity_title': 'Test Group 1',
                'entity_username': 'testgroup1'
            },
            {
                'scan_job_id': job.id,
                'source_id': source.id,
                'message_id': 2,
                'original_url': 'https://t.me/testchannel1',
                'normalized_url': 'https://t.me/testchannel1',
                'platform': 'telegram',
                'link_type': 'public_channel',
                'status': 'valid',
                'entity_title': 'Test Channel 1',
                'entity_username': 'testchannel1'
            },
            {
                'scan_job_id': job.id,
                'source_id': source.id,
                'message_id': 3,
                'original_url': 'https://t.me/+ABC123',
                'normalized_url': 'https://t.me/+ABC123',
                'platform': 'telegram',
                'link_type': 'private_invite',
                'status': 'valid',
                'entity_title': 'Test Invite'
            },
            {
                'scan_job_id': job.id,
                'source_id': source.id,
                'message_id': 4,
                'original_url': 'https://chat.whatsapp.com/XYZ789',
                'normalized_url': 'https://chat.whatsapp.com/XYZ789',
                'platform': 'whatsapp',
                'link_type': 'group',
                'status': 'valid'
            },
            {
                'scan_job_id': job.id,
                'source_id': source.id,
                'message_id': 5,
                'original_url': 'https://example.com',
                'normalized_url': 'https://example.com',
                'platform': 'other',
                'link_type': 'other',
                'status': 'excluded'
            }
        ]
        
        for link_data in test_links:
            await link_repo.create(**link_data)
        
        # Save statistics
        stats_data = {
            'telegram_count': 3,
            'whatsapp_count': 1,
            'group_count': 1,
            'channel_count': 1,
            'invite_count': 1,
            'personal_count': 0,
            'bot_count': 0,
            'duplicate_count': 0,
            'other_count': 1,
            'invalid_count': 0
        }
        
        await result_repo.save_statistics(job.id, stats_data)
        
        return job.id

    async def test_txt_export(self):
        """Test TXT export."""
        job_id = await self.setup_test_data()
        
        file_path = await TXTExporter.export(job_id)
        
        assert os.path.exists(file_path)
        
        # Read file and verify content
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        assert "TELEGRAM & WHATSAPP LINK EXTRACTOR" in content
        assert f"Job ID: {job_id}" in content
        assert "TELEGRAM GROUPS" in content
        assert "TELEGRAM CHANNELS" in content
        assert "TELEGRAM INVITES" in content
        assert "WHATSAPP GROUPS" in content
        assert "https://t.me/testgroup1" in content
        assert "https://t.me/testchannel1" in content
        assert "https://t.me/+ABC123" in content
        assert "https://chat.whatsapp.com/XYZ789" in content
        
        # Cleanup
        os.remove(file_path)

    async def test_csv_export(self):
        """Test CSV export."""
        job_id = await self.setup_test_data()
        
        file_path = await CSVExporter.export(job_id)
        
        assert os.path.exists(file_path)
        
        # Read CSV file
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        # Should have 5 rows (5 links)
        assert len(rows) == 5
        
        # Verify columns
        expected_columns = [
            'url', 'platform', 'link_type', 'status',
            'entity_title', 'entity_username', 'source_id', 'message_id',
            'created_at'
        ]
        for col in expected_columns:
            assert col in reader.fieldnames
        
        # Verify data
        telegram_rows = [r for r in rows if r['platform'] == 'telegram']
        whatsapp_rows = [r for r in rows if r['platform'] == 'whatsapp']
        other_rows = [r for r in rows if r['platform'] == 'other']
        
        assert len(telegram_rows) == 3
        assert len(whatsapp_rows) == 1
        assert len(other_rows) == 1
        
        # Cleanup
        os.remove(file_path)

    async def test_json_export(self):
        """Test JSON export."""
        job_id = await self.setup_test_data()
        
        file_path = await JSONExporter.export(job_id)
        
        assert os.path.exists(file_path)
        
        # Read JSON file
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Verify structure
        assert data['job_id'] == job_id
        assert 'statistics' in data
        assert 'links' in data
        
        # Verify links
        assert len(data['links']) == 5
        
        # Verify statistics
        assert data['statistics']['telegram_count'] == 3
        assert data['statistics']['whatsapp_count'] == 1
        
        # Cleanup
        os.remove(file_path)

    async def test_xlsx_export(self):
        """Test XLSX export."""
        job_id = await self.setup_test_data()
        
        file_path = await XLSXExporter.export(job_id)
        
        assert os.path.exists(file_path)
        
        # Verify file exists and has content
        assert Path(file_path).stat().st_size > 0
        
        # Check that openpyxl can open it
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        
        # Should have 2 sheets
        sheet_names = wb.sheetnames
        assert 'Results' in sheet_names
        assert 'Statistics' in sheet_names
        
        # Check Results sheet
        ws_results = wb['Results']
        assert ws_results.max_row > 1  # Has data
        
        # Check Statistics sheet
        ws_stats = wb['Statistics']
        assert ws_stats.max_row > 1  # Has data
        
        # Cleanup
        os.remove(file_path)

    async def test_export_cleanup(self):
        """Test that exported files are properly cleaned up."""
        job_id = await self.setup_test_data()
        
        # Export to TXT
        file_path = await TXTExporter.export(job_id)
        assert os.path.exists(file_path)
        os.remove(file_path)
        assert not os.path.exists(file_path)
        
        # Export to CSV
        file_path = await CSVExporter.export(job_id)
        assert os.path.exists(file_path)
        os.remove(file_path)
        assert not os.path.exists(file_path)
        
        # Export to JSON
        file_path = await JSONExporter.export(job_id)
        assert os.path.exists(file_path)
        os.remove(file_path)
        assert not os.path.exists(file_path)
        
        # Export to XLSX
        file_path = await XLSXExporter.export(job_id)
        assert os.path.exists(file_path)
        os.remove(file_path)
        assert not os.path.exists(file_path)